import os
import pytest
import ast
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playwright.sync_api import sync_playwright, Page
from utils.config import BROWSER_NAME, HEADLESS, DEFAULT_TIMEOUT, SLOW_MO, SCREENSHOT_DELAY
from utils.screenshot_helper import capture_failure_screenshot

# Global variables to store test results
test_results = {}
test_files_executed = set()

# REPORT GENERATION FUNCTIONALITY
# ============================================================================

def extract_test_descriptions(pyfile: str) -> dict:
    """
    Extracts docstrings from test functions and formats them with 'Verify' prefix.
    Returns dict: {test_function_name: description}
    """
    descriptions = {}
    try:
        with open(pyfile, "r", encoding="utf-8") as f:
            tree = ast.parse(f.read(), filename=pyfile)

        for node in tree.body:
            if isinstance(node, ast.FunctionDef) and node.name.startswith("test_"):
                docstring = ast.get_docstring(node)
                if docstring:
                    # Ensure description starts with "Verify"
                    description = docstring.strip()
                    if not description.lower().startswith("verify"):
                        description = f"Verify {description}"
                    descriptions[node.name] = description
                else:
                    # Fallback to cleaned test name with "Verify" prefix
                    clean_name = clean_test_name(node.name)
                    descriptions[node.name] = f"Verify {clean_name}"
    except Exception as e:
        print(f"Error extracting descriptions from {pyfile}: {e}")
    return descriptions

def clean_test_name(name: str) -> str:
    """Cleans a Python test function name for reporting."""
    if name.startswith("test_"):
        name = name[5:]
    parts = [part for part in name.split('_') if not part.isdigit()]
    clean_name = " ".join(parts)
    return clean_name.capitalize() if clean_name else name

def get_dynamic_test_data(test_name: str, test_file: str) -> dict:
    """
    Returns dynamically generated test data based on test name patterns.
    """
    # Extract module name from test file
    module_name = os.path.basename(test_file).replace("test_", "").replace(".py", "").replace("_", " ").title()
    
    # Common patterns for different types of tests
    if "login" in test_name.lower():
        return get_login_test_data(test_name)
    elif "signup" in test_name.lower() or "register" in test_name.lower():
        return get_signup_test_data(test_name)
    elif "email" in test_name.lower() and "verif" in test_name.lower():
        return get_email_verification_test_data(test_name)
    elif "password" in test_name.lower() and ("reset" in test_name.lower() or "forgot" in test_name.lower()):
        return get_password_reset_test_data(test_name)
    elif "agency" in test_name.lower():
        return get_agency_test_data(test_name)
    elif "jd" in test_name.lower() or "job" in test_name.lower():
        return get_jd_test_data(test_name)
    else:
        return get_generic_test_data(test_name, module_name)

def get_login_test_data(test_name: str) -> dict:
    """Returns specific test data for login tests."""
    if "successful" in test_name.lower():
        return {
            "pre_conditions": "User is on the login page and has valid credentials",
            "test_data": "Valid email and password",
            "test_steps": "1. Navigate to login page\n2. Enter valid credentials\n3. Click Sign in button",
            "expected_result": "User should be logged in successfully",
            "actual_result": "User logged in successfully"
        }
    elif "email" in test_name.lower() and "required" in test_name.lower():
        return {
            "pre_conditions": "User is on the login page",
            "test_data": "Empty email field",
            "test_steps": "1. Navigate to login page\n2. Leave email field empty\n3. Enter password\n4. Click Sign in",
            "expected_result": "System should display 'Email is required' error",
            "actual_result": "'Email is required' error message displayed"
        }
    elif "password" in test_name.lower() and "required" in test_name.lower():
        return {
            "pre_conditions": "User is on the login page",
            "test_data": "Empty password field",
            "test_steps": "1. Navigate to login page\n2. Enter email\n3. Leave password empty\n4. Click Sign in",
            "expected_result": "System should display 'Password is required' error",
            "actual_result": "'Password is required' error message displayed"
        }
    elif "invalid" in test_name.lower() and "email" in test_name.lower():
        return {
            "pre_conditions": "User is on the login page",
            "test_data": "Invalid email format",
            "test_steps": "1. Navigate to login page\n2. Enter invalid email\n3. Enter password\n4. Click Sign in",
            "expected_result": "System should display email format error",
            "actual_result": "Invalid email format error displayed"
        }
    elif "invalid" in test_name.lower() and "credential" in test_name.lower():
        return {
            "pre_conditions": "User is on the login page",
            "test_data": "Wrong credentials",
            "test_steps": "1. Navigate to login page\n2. Enter wrong credentials\n3. Click Sign in",
            "expected_result": "System should display invalid credentials error",
            "actual_result": "Invalid credentials error displayed"
        }
    else:
        return get_generic_test_data(test_name, "Login")

def get_signup_test_data(test_name: str) -> dict:
    """Returns specific test data for signup tests."""
    return {
        "pre_conditions": "User is on the signup page",
        "test_data": "Signup form data",
        "test_steps": "1. Navigate to signup page\n2. Fill required fields\n3. Submit form",
        "expected_result": "Signup process should work as expected",
        "actual_result": "Signup functionality verified"
    }

def get_email_verification_test_data(test_name: str) -> dict:
    """Returns specific test data for email verification tests."""
    return {
        "pre_conditions": "User has received verification email",
        "test_data": "OTP or verification code",
        "test_steps": "1. Open verification page\n2. Enter verification code\n3. Submit verification",
        "expected_result": "Email should be verified successfully",
        "actual_result": "Email verification completed"
    }

def get_password_reset_test_data(test_name: str) -> dict:
    """Returns specific test data for password reset tests."""
    return {
        "pre_conditions": "User needs to reset password",
        "test_data": "Email address for reset",
        "test_steps": "1. Navigate to forgot password\n2. Enter email\n3. Submit reset request",
        "expected_result": "Password reset process should work",
        "actual_result": "Password reset functionality verified"
    }

def get_agency_test_data(test_name: str) -> dict:
    """Returns specific test data for agency tests."""
    return {
        "pre_conditions": "User is logged in and has agency permissions",
        "test_data": "Agency related data",
        "test_steps": "1. Navigate to agency section\n2. Perform agency operations\n3. Verify results",
        "expected_result": "Agency functionality should work as expected",
        "actual_result": "Agency operations completed successfully"
    }

def get_generic_test_data(test_name: str, module_name: str) -> dict:
    """Returns generic test data for any test."""
    return {
        "pre_conditions": f"User is on the {module_name.lower()} page",
        "test_data": f"{module_name} test data",
        "test_steps": f"1. Navigate to {module_name.lower()} page\n2. Perform required actions\n3. Verify expected behavior",
        "expected_result": f"{module_name} functionality should work as expected",
        "actual_result": f"{module_name} functionality verified"
    }

def generate_excel_report(test_file: str, test_descriptions: dict, results: dict) -> None:
    """
    Creates an Excel file with detailed test report.
    """
    # Generate report filename
    base_name = os.path.basename(test_file).replace(".py", "")
    report_filename = f"reports/{base_name}_report.xlsx"
    
    # Ensure reports directory exists
    os.makedirs("reports", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    # Truncate title to 31 characters to avoid Excel warning
    title = f"{base_name.replace('_', ' ').title()} Test Report"
    ws.title = title[:31] if len(title) > 31 else title

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Header row
    headers = [
        "Serial No.", "Test Case Id.", "Test Objective", "Pre-Conditions",
        "Test Data", "Test Steps", "Expected Result", "Actual Result", "Test Status"
    ]
    ws.append(headers)
    
    # Apply header styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Data rows
    for idx, (test_name, desc) in enumerate(test_descriptions.items(), start=1):
        status = results.get(test_name, "Not Run")
        #tc_id = f"TC_{base_name.upper()}_{idx:03d}"
        tc_id = f"TC_{idx:03d}"
        test_data = get_dynamic_test_data(test_name, test_file)
        
        # Determine actual result based on test status
        if status == "PASSED":
            actual_result = test_data["actual_result"]
        elif status == "FAILED":
            actual_result = "Test failed - actual behavior did not match expected result"
        elif status == "SKIPPED":
            actual_result = "Test was skipped during execution"
        else:
            actual_result = "Test not executed"
        
        row_data = [
            f"{idx:02d}", tc_id, desc, test_data["pre_conditions"],
            test_data["test_data"], test_data["test_steps"], 
            test_data["expected_result"], actual_result, status
        ]
        
        ws.append(row_data)
        
        # Apply styling to data rows
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.border = thin_border
            if col_num == 1:  # Serial No.
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
            
            # Color code test status
            if col_num == 9:  # Test Status column
                if status == "PASSED":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif status == "FAILED":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif status == "SKIPPED":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Adjust column widths
    column_widths = [10, 15, 40, 30, 30, 50, 40, 40, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Set row heights
    for row in range(2, len(test_descriptions) + 2):
        ws.row_dimensions[row].height = 80

    wb.save(report_filename)
    print(f"Test report generated: {report_filename}")

# BROWSER AND PAGE FIXTURES
# ============================================================================

@pytest.fixture(scope="session")
def browser():
    """Session-scoped browser fixture with configuration from config.py"""
    with sync_playwright() as p:
        if BROWSER_NAME == "chromium":
            browser = p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO)
        elif BROWSER_NAME == "firefox":
            browser = p.firefox.launch(headless=HEADLESS, slow_mo=SLOW_MO)
        elif BROWSER_NAME == "webkit":
            browser = p.webkit.launch(headless=HEADLESS, slow_mo=SLOW_MO)
        else:
            browser = p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO)
            
        yield browser
        browser.close()

@pytest.fixture
def context(browser, request):
    """Browser context fixture with tracing support."""
    context = browser.new_context()
    tracing_enabled = request.config.getoption("--tracing") if hasattr(request.config, 'getoption') else False
    if tracing_enabled:
        context.tracing.start(screenshots=True, snapshots=True, sources=True)
    yield context
    if tracing_enabled:
        trace_dir = os.path.join(os.getcwd(), "traces")
        os.makedirs(trace_dir, exist_ok=True)
        trace_file = os.path.join(trace_dir, f"trace_{request.node.name}.zip")
        context.tracing.stop(path=trace_file)
    context.close()

@pytest.fixture
def page(context):
    """Page fixture using context, with configured timeout."""
    page = context.new_page()
    page.set_default_timeout(DEFAULT_TIMEOUT)
    yield page
    page.close()

# GLOBAL UTILITY FUNCTIONS
# ============================================================================

def wait_for_action_completion(page: Page, action_type: str = "general"):
    """
    Global utility function to wait for various actions to complete.
    
    Args:
        page: Playwright page object
        action_type: Type of action performed (signup, login, verify, etc.)
    """
    # Always wait for network to be idle first
    page.wait_for_load_state("networkidle", timeout=15000)
    
    # Add specific waits based on action type
    if action_type == "signup":
        page.wait_for_timeout(3000)  # 3 seconds for signup success/error messages
    elif action_type == "login":
        page.wait_for_timeout(2000)  # 2 seconds for login processing ONLY
    elif action_type == "verify" or action_type == "otp":
        page.wait_for_timeout(4000)  # 4 seconds for verification processing
    elif action_type == "save" or action_type == "update":
        page.wait_for_timeout(2500)  # 2.5 seconds for save operations
        # Additional wait for success/error messages
        page.wait_for_timeout(1500)  # Extra 1.5 seconds for messages
    elif action_type == "navigation":
        page.wait_for_timeout(1500)  # 1.5 seconds for navigation
    elif action_type == "modal":
        page.wait_for_timeout(3000)  # 3 seconds specifically for modal appearance
    else:
        page.wait_for_timeout(2000)  # Default 2 seconds

def wait_for_modal_or_content(page: Page, timeout: int = 10000):
    """
    Global utility to wait for modals or dynamic content to appear after login.
    """
    try:
        # Wait for any dynamic content to load
        page.wait_for_load_state("networkidle", timeout=timeout)
        page.wait_for_timeout(3000)  # Additional 3 seconds for modals
        
        # Try to detect if any modal or overlay is present
        modal_selectors = [
            '[role="dialog"]',
            '.modal',
            '[data-modal]',
            '.overlay',
            '[aria-modal="true"]'
        ]
        
        for selector in modal_selectors:
            try:
                if page.locator(selector).count() > 0:
                    page.wait_for_timeout(1000)  # Extra wait if modal detected
                    break
            except:
                continue
                
    except Exception as e:
        print(f"⚠️  Modal wait completed with timeout: {e}")

def capture_immediate_screenshot(page: Page, test_name: str, description: str = "failure"):
    """
    Immediately capture screenshot for failed assertions.
    This function is called right when an assertion fails.
    """
    try:
        timestamp = datetime.now().strftime("%d-%m-%Y_%H.%M.%S")
        
        # Get the test file from the call stack to determine correct folder
        import inspect
        frame = inspect.currentframe()
        test_file = ""
        while frame:
            frame_info = inspect.getframeinfo(frame)
            if frame_info.filename and 'test_' in frame_info.filename:
                test_file = frame_info.filename
                break
            frame = frame.f_back
        
        # Determine screenshot folder based on test file name
        if "test_login.py" in test_file:
            folder = "login_screenshots"
        elif "test_signup.py" in test_file:
            folder = "signup_screenshots"
        elif "test_agency.py" in test_file:
            folder = "agency_screenshots"
        elif "test_email" in test_file:
            folder = "email_verify_screenshots"
        elif "test_reset" in test_file:
            folder = "reset_pass_screenshots"
        elif "test_jd.py" in test_file:
            folder = "jd_screenshots"
        else:
            folder = "general_screenshots"
        
        # Create directory
        screenshot_dir = os.path.join("screenshots", folder)
        os.makedirs(screenshot_dir, exist_ok=True)
        
        # Generate filename in format: TC_XX_DD-MM-YYYY_HH.MM.SS.png
        clean_test_name = test_name.replace("test_", "").replace("_", "_")
        filename = f"{clean_test_name}_{description}_{timestamp}.png"
        filepath = os.path.join(screenshot_dir, filename)
        
        # Take screenshot immediately
        page.screenshot(path=filepath)
        print(f"📸 Immediate screenshot saved: {filepath}")
        return filepath
        
    except Exception as e:
        print(f"⚠️  Failed to capture immediate screenshot: {e}")
        return None

# SCREENSHOT HOOKS AND FIXTURES
# ============================================================================

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Pytest hook to capture test execution results and screenshots for failures."""
    outcome = yield
    rep = outcome.get_result()
    setattr(item, f"rep_{rep.when}", rep)
    
    # Capture test results for report generation
    if rep.when == "call":  # Only capture the actual test execution result
        test_file = str(item.fspath)
        test_name = item.name
        
        # Store the test file being executed
        test_files_executed.add(test_file)
        
        # Store test result
        if test_file not in test_results:
            test_results[test_file] = {}
        
        if rep.passed:
            test_results[test_file][test_name] = "PASSED"
        elif rep.failed:
            test_results[test_file][test_name] = "FAILED"
            
            # Screenshot capture is now handled by enhanced assertions
            # No need to capture screenshots here to avoid duplicates
                
        elif rep.skipped:
            test_results[test_file][test_name] = "SKIPPED"

def pytest_sessionfinish(session, exitstatus):
    """Generate reports after all tests complete."""
    for test_file in test_files_executed:
        if test_file in test_results and test_results[test_file]:
            # Extract test descriptions from the file
            test_descriptions = extract_test_descriptions(test_file)
            
            # Generate Excel report
            generate_excel_report(test_file, test_descriptions, test_results[test_file])
    
    # Clear the results for next run
    test_results.clear()
    test_files_executed.clear()

# JD TEST FIXTURES AND DATA MANAGEMENT
# ============================================================================

@pytest.fixture(scope="module")
def jd_test_data():
    """
    Module-scoped fixture providing JD test data for all tests in a module.
    Creates fresh data for each test module to ensure isolation.
    """
    from random_values_generator.random_jd_data import (
        generate_complete_jd_data, 
        generate_minimal_jd_data,
        generate_multiple_jd_data,
        generate_search_test_data,
        generate_filter_test_data,
        generate_invalid_jd_data_cases
    )
    
    return {
        'complete_jd': generate_complete_jd_data(),
        'minimal_jd': generate_minimal_jd_data(),
        'multiple_jds': generate_multiple_jd_data(5),
        'search_data': generate_search_test_data(),
        'filter_data': generate_filter_test_data(),
        'invalid_cases': generate_invalid_jd_data_cases()
    }

@pytest.fixture(scope="function")
def fresh_jd_data():
    """
    Function-scoped fixture providing fresh JD data for each test.
    Use this when tests need unique data that won't conflict.
    """
    from random_values_generator.random_jd_data import generate_complete_jd_data
    return generate_complete_jd_data()

@pytest.fixture(scope="function")
def jd_bulk_data():
    """
    Function-scoped fixture providing bulk JD data for performance testing.
    """
    from random_values_generator.random_jd_data import generate_bulk_jd_data
    return generate_bulk_jd_data(20)

@pytest.fixture(scope="function")
def jd_validation_data():
    """
    Function-scoped fixture providing validation test cases for JD testing.
    """
    from random_values_generator.random_jd_data import (
        generate_invalid_jd_data_cases,
        generate_edge_case_jd_data
    )
    
    return {
        'invalid_cases': generate_invalid_jd_data_cases(),
        'edge_cases': generate_edge_case_jd_data()
    }

@pytest.fixture(scope="session")
def jd_cleanup_tracker():
    """
    Session-scoped fixture to track created JDs for cleanup.
    Maintains a list of JD IDs that need to be cleaned up after tests.
    """
    created_jds = []
    yield created_jds
    
    # Cleanup logic would go here if we had API access
    # For now, we'll just log what needs cleanup
    if created_jds:
        print(f"\n🧹 JDs created during testing (for manual cleanup): {len(created_jds)}")
        for jd_id in created_jds:
            print(f"   - JD ID: {jd_id}")

@pytest.fixture(scope="function")
def jd_test_context(page, jd_cleanup_tracker):
    """
    Function-scoped fixture providing JD test context with cleanup tracking.
    Use this fixture when you need to track JDs created during testing.
    """
    context = {
        'page': page,
        'created_jds': [],
        'cleanup_tracker': jd_cleanup_tracker
    }
    
    yield context
    
    # Add created JDs to session cleanup tracker
    jd_cleanup_tracker.extend(context['created_jds'])

@pytest.fixture(scope="function")
def jd_search_context():
    """
    Function-scoped fixture providing search-specific test context.
    """
    from random_values_generator.random_jd_data import generate_search_test_data
    
    search_data = generate_search_test_data()
    return {
        'search_terms': [
            search_data['searchable_jd'].position_title.split()[0],
            search_data['searchable_jd'].company.split()[0],
            search_data['unique_jd'].department,
            "NonExistentTerm12345"  # For no results testing
        ],
        'expected_results': {
            'searchable': search_data['searchable_jd'],
            'unique': search_data['unique_jd'],
            'common': search_data['common_jd']
        }
    }

@pytest.fixture(scope="function")
def jd_filter_context():
    """
    Function-scoped fixture providing filter-specific test context.
    """
    from random_values_generator.random_jd_data import generate_filter_test_data
    
    filter_data = generate_filter_test_data()
    return {
        'filter_combinations': [
            {'work_style': 'Remote'},
            {'hiring_status': 'Open'},
            {'employment_type': 'Full-time'},
            {'work_style': 'Remote', 'hiring_status': 'Open'},
            {'work_style': 'Hybrid', 'employment_type': 'Part-time'}
        ],
        'test_jds': filter_data
    }

@pytest.fixture(scope="function")
def jd_pagination_context():
    """
    Function-scoped fixture providing pagination test context.
    """
    return {
        'page_sizes': [5, 10, 20, 50],
        'navigation_patterns': [
            'first_to_last',
            'last_to_first', 
            'random_pages',
            'sequential_forward',
            'sequential_backward'
        ]
    }

# JD DATA MANAGEMENT UTILITIES
# ============================================================================

def get_jd_test_data(test_name: str) -> dict:
    """
    Returns JD-specific test data based on test name patterns.
    """
    if "create" in test_name.lower() or "add" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and on JD management page",
            "test_data": "Valid JD creation data with all required fields",
            "test_steps": "1. Click Add JD button\n2. Fill required fields\n3. Save JD",
            "expected_result": "JD should be created successfully with success message",
            "actual_result": "JD created successfully"
        }
    elif "edit" in test_name.lower() or "update" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and JD exists in the system",
            "test_data": "Updated JD information",
            "test_steps": "1. Navigate to JD list\n2. Click edit on existing JD\n3. Modify fields\n4. Save changes",
            "expected_result": "JD should be updated successfully",
            "actual_result": "JD updated successfully"
        }
    elif "delete" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and JD exists in the system",
            "test_data": "Existing JD to delete",
            "test_steps": "1. Navigate to JD list\n2. Click delete on JD\n3. Confirm deletion",
            "expected_result": "JD should be deleted successfully",
            "actual_result": "JD deleted successfully"
        }
    elif "search" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and JDs exist in the system",
            "test_data": "Search terms for JD lookup",
            "test_steps": "1. Navigate to JD list\n2. Enter search term\n3. Execute search",
            "expected_result": "Relevant JDs should be displayed in search results",
            "actual_result": "Search results displayed correctly"
        }
    elif "filter" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and JDs exist with different attributes",
            "test_data": "Filter criteria (work style, status, etc.)",
            "test_steps": "1. Navigate to JD list\n2. Open filters\n3. Apply filter criteria",
            "expected_result": "JDs matching filter criteria should be displayed",
            "actual_result": "Filtered results displayed correctly"
        }
    elif "validation" in test_name.lower() or "required" in test_name.lower():
        return {
            "pre_conditions": "User is on JD creation/edit form",
            "test_data": "Invalid or missing field data",
            "test_steps": "1. Open JD form\n2. Leave required fields empty or enter invalid data\n3. Attempt to save",
            "expected_result": "Appropriate validation error messages should be displayed",
            "actual_result": "Validation errors displayed correctly"
        }
    elif "upload" in test_name.lower() or "file" in test_name.lower():
        return {
            "pre_conditions": "User is on JD page with file upload capability",
            "test_data": "Test files (valid and invalid formats)",
            "test_steps": "1. Click upload file button\n2. Select test file\n3. Upload file",
            "expected_result": "File should be processed according to format validation rules",
            "actual_result": "File upload processed correctly"
        }
    elif "pagination" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and multiple JDs exist (more than page size)",
            "test_data": "Large dataset of JDs",
            "test_steps": "1. Navigate to JD list\n2. Navigate through pages\n3. Verify pagination controls",
            "expected_result": "Pagination should work correctly with proper navigation",
            "actual_result": "Pagination functionality verified"
        }
    elif "bulk" in test_name.lower():
        return {
            "pre_conditions": "User is logged in and multiple JDs exist",
            "test_data": "Multiple JDs for bulk operations",
            "test_steps": "1. Navigate to JD list\n2. Select multiple JDs\n3. Perform bulk operation",
            "expected_result": "Bulk operation should be applied to all selected JDs",
            "actual_result": "Bulk operation completed successfully"
        }
    else:
        return {
            "pre_conditions": "User is logged in and on JD management page",
            "test_data": "JD test data",
            "test_steps": "1. Navigate to JD page\n2. Perform JD operations\n3. Verify results",
            "expected_result": "JD functionality should work as expected",
            "actual_result": "JD functionality verified"
        }

# Note: Screenshots are automatically captured for failed tests in pytest_runtest_makereport hook above